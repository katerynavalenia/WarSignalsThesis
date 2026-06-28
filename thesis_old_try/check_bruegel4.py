import openpyxl

wb = openpyxl.load_workbook(
    r"C:\Users\A00010311\Downloads\Master Thesis\Europe-Central-Asia_aggregated_data_up_to_week_of-2026-06-06.xlsx",
    read_only=True, data_only=True
)
ws = wb.active
# Read first 3 rows and many columns to understand structure
print("=== First 3 rows, up to 50 cols ===")
for i, row in enumerate(ws.iter_rows(max_row=3, values_only=True)):
    print(f"Row {i}: {list(row)[:50]}")
    
print("\n=== Count rows and cols by iterating ===")
col_count = 0
row_count = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        col_count = len([x for x in row if x is not None])
    row_count += 1
    if i > 300:
        print(f"... stopping at row {i}")
        break
print(f"Row count (up to 300): {row_count}, Non-null cols in row 0: {col_count}")
wb.close()
