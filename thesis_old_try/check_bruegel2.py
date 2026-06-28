import pandas as pd
import warnings
warnings.filterwarnings("ignore")
import openpyxl

wb = openpyxl.load_workbook(
    r"C:\Users\A00010311\Downloads\Master Thesis\Europe-Central-Asia_aggregated_data_up_to_week_of-2026-06-06.xlsx",
    read_only=True, data_only=True
)
print("Sheets:", wb.sheetnames)
ws = wb.active
print("Max row:", ws.max_row, "Max col:", ws.max_column)
rows = list(ws.iter_rows(max_row=6, values_only=True))
for i, r in enumerate(rows):
    print(f"Row {i}: {r[:15]}")
wb.close()
