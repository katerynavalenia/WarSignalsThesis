import openpyxl

wb = openpyxl.load_workbook(
    r"C:\Users\A00010311\Downloads\Master Thesis\Europe-Central-Asia_aggregated_data_up_to_week_of-2026-06-06.xlsx",
    read_only=True, data_only=True
)
ws = wb.active
print(f"max_row attr: {ws.max_row}, max_col attr: {ws.max_column}")

# Read ALL rows in ONE pass
all_rows = []
for row in ws.iter_rows(max_row=5000, max_col=300, values_only=True):
    all_rows.append(row)

print(f"Total rows found: {len(all_rows)}")
if all_rows:
    total_cols = max(len(r) for r in all_rows)
    print(f"Max cols across all rows: {total_cols}")
    for i, r in enumerate(all_rows[:10]):
        non_none = [x for x in r if x is not None]
        print(f"  Row {i}: {len(r)} total cells, {len(non_none)} non-null | sample: {non_none[:8]}")
wb.close()
